# api/main.py

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel
import os
from dotenv import load_dotenv
import google.generativeai as genai
import chromadb
from typing import List, Dict, Any
from pathlib import Path

# ✅ Use your existing indexer (the one that uses SheetParser internally)
from utils.index_data import index_spreadsheet_data

import json
from openpyxl import load_workbook

# --- Pydantic Data Models ---


class SearchRequest(BaseModel):
    query: str


class SearchResult(BaseModel):
    query: str
    result: str
    context: List[Dict[str, Any]]


# --- CONFIGURATION & INITIALIZATION ---
load_dotenv()
API_KEY = os.getenv("LLM_API_KEY")
if not API_KEY:
    # Do not raise here. Let startup handle missing key (or run in degraded mode).
    print("⚠️ LLM_API_KEY not found in environment. Gemini features will be disabled until the key is set.")


CHROMA_PATH = "chroma_db"
COLLECTION_NAME = "spreadsheet_concepts"
EMBEDDING_MODEL = "models/text-embedding-004"
GENERATIVE_MODEL = "gemini-2.5-flash"
K_RESULTS = 15

# Global variables
LLM_client = None
chroma_client = None
collection = None
collection_count = 0  # number of concepts currently indexed

DATA_DIR = Path(os.getcwd()) / "data"

# Initialize FastAPI app
app = FastAPI(title="Semantic Spreadsheet Search API")

# CORS for React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # adjust as needed
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- HELPER: CHECK IF ANY CURRENT SPREADSHEETS EXIST ---


def current_dataset_available() -> bool:
    """Returns True if there is at least one .xlsx file in ./data."""
    if not DATA_DIR.exists():
        return False
    return any(DATA_DIR.glob("*.xlsx"))


def clear_index(reset_data_dir: bool = False):
    """
    Drop the Chroma collection and optionally delete all files in DATA_DIR.
    Used when we want a clean slate (e.g. new browser session).
    """
    global collection, collection_count, chroma_client

    # Delete Chroma collection
    if chroma_client:
        try:
            chroma_client.delete_collection(name=COLLECTION_NAME)
            print("🧹 Deleted Chroma collection:", COLLECTION_NAME)
        except Exception as e:
            print("⚠️ Could not delete Chroma collection:", e)

    collection = None
    collection_count = 0

    # Optionally wipe data directory
    if reset_data_dir and DATA_DIR.exists():
        for f in DATA_DIR.glob("*"):
            if f.is_file():
                try:
                    f.unlink()
                except Exception as e:
                    print("⚠️ Could not delete file:", f, e)


# --- STARTUP INITIALIZATION ---


@app.on_event("startup")
def initialize_clients():
    global LLM_client, chroma_client
    # Initialize Gemini API (for query/synthesis)
    try:
        genai.configure(api_key=API_KEY)
        LLM_client = genai
        print("✅ Gemini API configured successfully.")
    except Exception as e:
        print(f"❌ Error configuring Gemini API: {e}")

    # Initialize ChromaDB client
    try:
        chroma_client = chromadb.PersistentClient(path=CHROMA_PATH)
        # Try to load existing collection, if any
        try:
            load_collection_status()
        except Exception:
            print(
                "ChromaDB client initialized. Collection will be created upon first indexing."
            )
    except Exception as e:
        print(f"❌ Error initializing ChromaDB client: {e}")


def load_collection_status():
    """
    Helper to check and set global collection status from ChromaDB.
    Called from startup and /status endpoint.
    """
    global collection, collection_count
    if chroma_client:
        try:
            collection = chroma_client.get_collection(name=COLLECTION_NAME)
            collection_count = collection.count()
            print(f"📚 ChromaDB collection loaded. Indexed concepts: {collection_count}")
        except Exception:
            collection = None
            collection_count = 0
            print("ℹ️ ChromaDB collection not found yet.")
    else:
        collection = None
        collection_count = 0


def build_example_queries(max_examples: int = 5) -> List[str]:
    """
    Build sheet-aware example queries from the existing ChromaDB collection.

    Uses metadata fields:
      - sheet
      - header
    """
    if collection is None or collection_count == 0:
        return []

    try:
        # Pull a sample of metadata rows (no need to fetch the whole DB)
        limit = min(collection_count, 50)
        res = collection.get(limit=limit, include=["metadatas"])
    except Exception as e:
        print(f"⚠️ Failed to build example queries from collection: {e}")
        return []

    metadatas = res.get("metadatas") or []
    seen_keys = set()
    headers = []

    for meta in metadatas:
        if not isinstance(meta, dict):
            continue
        header = meta.get("header")
        sheet = meta.get("sheet")
        if not header:
            continue
        key = (sheet or "", header)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        headers.append((sheet, header))

    examples: List[str] = []
    for sheet, header in headers[:max_examples]:
        header_str = str(header).strip()
        sheet_str = str(sheet).strip() if sheet else ""

        if sheet_str:
            q = f"What is {header_str} in sheet {sheet_str}?"
        else:
            q = f"What is {header_str}?"
        examples.append(q)

    return examples


# --- CORE HELPERS ---


def embed_query(query: str, model: str = EMBEDDING_MODEL):
    """Generates the vector embedding for the user's search query."""
    if not LLM_client:
        raise HTTPException(status_code=500, detail="Gemini API not initialized.")
    try:
        response = LLM_client.embed_content(model=model, content=query)
        # For single input: response["embedding"]
        return response["embedding"]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Embedding error: {e}")


def query_vector_db(query_embedding):
    """Performs the vector search against the ChromaDB collection."""
    if collection is None or collection_count == 0:
        raise HTTPException(
            status_code=503,
            detail="Database not initialized or empty. Please upload and index data first.",
        )

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=K_RESULTS,
        include=["documents", "metadatas"],
    )
    return results["documents"][0], results["metadatas"][0]


def synthesize_answer(
    query: str, retrieved_context: list, retrieved_metadata: list
) -> str:
    """
    Uses the Gemini model to generate a natural language answer based on retrieved context.
    """
    formatted_context = []
    for doc, meta in zip(retrieved_context, retrieved_metadata):
        formatted_context.append(
            f"* Source: **{meta.get('sheet', 'N/A')}** | "
            f"Metric: **{meta.get('header', 'N/A')}** | "
            f"Context: {doc}"
        )

    context_str = "\n".join(formatted_context)

    full_prompt = (
        "You are a semantic search assistant for large spreadsheets. Your task is to use the "
        "provided context snippets (retrieved from a vector database) to answer the user's "
        "natural language query about the spreadsheet data. "
        "**Crucially, if the query requires a comparison (e.g., 'best', 'highest', 'who') or a simple calculation "
        "(e.g., percentage, total), you MUST perform that calculation/ranking based on the provided context before "
        "generating the final answer.** "
        "Only use the information provided in the context. If the context does not contain the answer, state that "
        "you cannot find the relevant information. Use relevant symbols like percentage(%) wherever required. "
        "Structure your answer clearly by grouping information by source sheet.\n\n"
        f"User Query: {query}\n\n"
        f"--- CONTEXT SNIPPETS ---\n"
        f"{context_str}\n\n"
        "Please synthesize the answer to the User Query using ONLY the context provided above. "
        "Provide the specific sheet and metric name for each piece of information."
    )

    try:
        model = genai.GenerativeModel(model_name=GENERATIVE_MODEL)
        response = model.generate_content(contents=[full_prompt])
        return response.text
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Synthesis error: {e}")


# --- INDEXING ENDPOINT (MULTIPART UPLOAD) ---


@app.post("/index")
async def index_data(files: List[UploadFile] = File(...)):
  """
  Upload Excel files via multipart/form-data, save them into ./data,
  then call your existing index_spreadsheet_data() (which uses SheetParser)
  to build the ChromaDB index.

  🔹 We do NOT modify the workbook content here.
  """
  global collection_count

  if not files:
      raise HTTPException(status_code=400, detail="No files uploaded.")

  # Ensure data directory exists
  DATA_DIR.mkdir(exist_ok=True)

  # Clear existing files and index (fresh dataset for this upload)
  clear_index(reset_data_dir=True)

  # Save new uploads exactly as they are
  for upload in files:
      dest = DATA_DIR / upload.filename
      with dest.open("wb") as f:
          f.write(await upload.read())
      print(f"📁 Saved uploaded file to: {dest}")

  # Now call your proven indexer, which reads from ./data and builds the collection
  try:
      print("🚀 Starting indexing via index_data.index_spreadsheet_data() ...")
      index_spreadsheet_data()

      # After indexing, refresh collection & count
      load_collection_status()

      if collection_count == 0:
          return {
              "status": "error",
              "indexed_concepts": 0,
              "message": "Indexing finished but no concepts were indexed.",
              "example_queries": [],
          }

      # 🔹 Build curated examples based on the actual indexed sheets/metrics
      example_queries = build_example_queries()

      return {
          "status": "success",
          "indexed_concepts": collection_count,
          "message": f"Files successfully uploaded and indexed. Total concepts: {collection_count}.",
          "example_queries": example_queries,
      }
  except Exception as e:
      print(f"❌ Indexing failed: {e}")
      raise HTTPException(status_code=500, detail=f"Indexing failed: {e}")


# --- RESET ENDPOINT (CALLED ON APP LOAD) ---


@app.post("/reset")
def reset_all():
    """
    Completely clear current dataset + index.
    Intended to be called once when the frontend app loads/refreshes.
    """
    clear_index(reset_data_dir=True)
    return {
        "status": "reset",
        "indexed_concepts": 0,
        "example_queries": [],
    }


# --- FILE LIST & DOWNLOAD / PREVIEW ENDPOINTS ---


@app.get("/files")
def list_files():
    """
    List the currently uploaded Excel files in ./data.
    Used by the frontend to show download links.
    """
    if not current_dataset_available():
        return {"files": []}

    files = []
    for p in DATA_DIR.glob("*.xlsx"):
        try:
            stat = p.stat()
            files.append(
                {
                    "name": p.name,
                    "size_bytes": stat.st_size,
                }
            )
        except Exception:
            files.append({"name": p.name, "size_bytes": None})

    return {"files": files}


@app.get("/files/{filename}")
def download_file(filename: str):
    """
    Download the Excel file from ./data.

    🔹 No edits are ever applied now – this is exactly what was uploaded.
    """
    target = DATA_DIR / filename
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found.")

    return FileResponse(
        path=str(target),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        filename=filename,
    )


# ---------- SIMPLE HTML PREVIEW (optional / legacy) ----------


@app.get("/preview/{filename}", response_class=HTMLResponse)
def preview_file(filename: str):
    """
    Simple HTML preview of the Excel file for inline 'View Sheet'.
    This does NOT affect what gets downloaded.
    """
    file_path = DATA_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")

    try:
        # data_only=True → shows cached values if present
        wb_values = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read workbook: {e}")

    sheets_val = wb_values.worksheets
    if not sheets_val:
        raise HTTPException(status_code=500, detail="Workbook has no sheets.")

    # Preview limits
    max_rows = 40
    max_cols = 15
    max_sheets = 6

    html_parts = [
        "<html><head><meta charset='utf-8' />",
        "<style>",
        "body { font-family: system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; padding: 12px; }",
        "h2 { margin: 0 0 12px; font-size: 16px; }",
        "h3 { margin: 16px 0 8px; font-size: 14px; }",
        "table { border-collapse: collapse; font-size: 12px; width: 100%; margin-bottom: 12px; }",
        "th, td { border: 1px solid #e5e7eb; padding: 4px 6px; text-align: left; }",
        "th { background: #f3f4f6; }",
        "tr:nth-child(even) td { background: #f9fafb; }",
        "</style></head><body>",
        f"<h2>{filename}</h2>",
    ]

    for idx, sheet_val in enumerate(sheets_val[:max_sheets], start=1):
        sheet_name = sheet_val.title
        html_parts.append(f"<h3>Sheet {idx}: {sheet_name}</h3>")
        html_parts.append("<table>")

        for i, row in enumerate(sheet_val.iter_rows(values_only=True), start=1):
            if i > max_rows:
                break
            html_parts.append("<tr>")
            for j, value in enumerate(row, start=1):
                if j > max_cols:
                    break
                display = "" if value is None else str(value)
                tag = "th" if i == 1 else "td"
                html_parts.append(f"<{tag}>{display}</{tag}>")
            html_parts.append("</tr>")

        html_parts.append("</table>")

    html_parts.append("</body></html>")
    html = "".join(html_parts)
    return HTMLResponse(content=html)


# ---------- SIMPLE JSON PREVIEW (for React, optional) ----------


@app.get("/preview-json/{filename}")
def preview_json(filename: str):
    """
    Very simple JSON representation of the workbook.

    This is only for UI preview. It does NOT affect the actual workbook
    that is returned by /files/{filename}.
    """
    if not current_dataset_available():
        raise HTTPException(
            status_code=503,
            detail="No spreadsheets uploaded. Please upload and index spreadsheets first.",
        )

    target = DATA_DIR / filename
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found.")

    try:
        wb = load_workbook(str(target), data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read workbook for preview: {e}",
        )

    sheets_payload: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        sheet_rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            row_vals = []
            for v in row:
                if v is None:
                    row_vals.append("")
                else:
                    row_vals.append(str(v))
            sheet_rows.append(row_vals)

        sheets_payload.append(
            {
                "name": ws.title,
                "rows": sheet_rows,
            }
        )

    return {"sheets": sheets_payload}


# --- STATUS & SEARCH ENDPOINTS ---


@app.get("/status")
def get_status():
    """
    Returns current indexing status and example queries.
    If there are no spreadsheets currently uploaded, we explicitly
    report 0 concepts and no examples so the UI disables search.
    """
    if not current_dataset_available():
        # Just in case, keep index cleared if there's no dataset
        clear_index(reset_data_dir=False)
        return {
            "status": "no_data",
            "indexed_concepts": 0,
            "example_queries": [],
        }

    # Dataset exists → report actual index status
    load_collection_status()
    example_queries = build_example_queries()
    return {
        "status": "ok",
        "indexed_concepts": collection_count,
        "example_queries": example_queries,
    }


@app.post("/search", response_model=SearchResult)
def search_data(request: SearchRequest):
    """
    Performs the full RAG pipeline:
    1. Embed query
    2. Query ChromaDB
    3. Synthesize answer using Gemini
    """
    if not current_dataset_available():
        raise HTTPException(
            status_code=503,
            detail="No spreadsheets uploaded. Please upload and index spreadsheets first.",
        )

    query = request.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query cannot be empty.")

    # 1. Embed Query
    query_vector = embed_query(query)

    # 2. Query Vector DB
    retrieved_docs, retrieved_metadatas = query_vector_db(query_vector)

    # 3. Synthesize Answer
    final_answer = synthesize_answer(query, retrieved_docs, retrieved_metadatas)

    # 4. Format context for frontend
    context_list = []
    for doc, meta in zip(retrieved_docs, retrieved_metadatas):
        context_list.append(
            {
                "sheet": meta.get("sheet", "N/A"),
                "metric": meta.get("header", "N/A"),
                "snippet": doc,
            }
        )

    return SearchResult(query=query, result=final_answer, context=context_list)
