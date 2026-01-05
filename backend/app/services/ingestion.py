# backend/app/services/ingestion.py
import pandas as pd
from openpyxl import load_workbook
from typing import Iterator, List, Any, Optional

class SpreadsheetRow:
    def __init__(self, sheet_name: str, row_index: int, row_header: str, values: List[Any], context_tags: List[str]):
        self.sheet_name = sheet_name
        self.row_index = row_index
        self.row_header = str(row_header).strip() if row_header else "Unknown"
        self.values = values
        self.context_tags = context_tags
        self.semantic_text = self._generate_semantic_text()

    def _generate_semantic_text(self) -> str:
        """
        Creates a rich, self-describing text representation for the AI.
        FIX: Maps Headers to Values and removes the [:4] truncation limit.
        """
        content_parts = []
        
        for i, val in enumerate(self.values):
            # Skip empty cells to save tokens and reduce noise
            if val is None or str(val).strip() == "":
                continue
            
            # Match value with its column header (safeguard against index errors)
            # context_tags[0] corresponds to values[0] because we stripped the row label column earlier
            header = self.context_tags[i] if i < len(self.context_tags) else f"Col_{i}"
            
            # Format: "Header=Value" (e.g., "Safety Stock=500")
            content_parts.append(f"{header}={str(val)}")
        
        # Join all parts. No truncation limit.
        data_str = "; ".join(content_parts)

        return (
            f"Sheet: {self.sheet_name} | "
            f"Row_Item: {self.row_header} | "
            f"Data: {data_str}"
        )

    def to_metadata(self):
        return {
            "sheet": self.sheet_name,
            "row_idx": self.row_index,
            "header": self.row_header
        }

class StreamingSheetParser:
    """
    Parses Excel files iteratively using openpyxl read_only mode.
    Optimized for large files on limited CPU/RAM.
    """
    def __init__(self, file_path: str):
        self.file_path = file_path

    def process_generator(self) -> Iterator[SpreadsheetRow]:
        """Yields SpreadsheetRow objects one by one with empty-row safeguards."""
        print(f"📄 Streaming parsing: {self.file_path}")
        
        try:
            # read_only=True and data_only=True are essential for performance and ignoring formulas
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                # Skip sheets that are usually just for display (dashboards/charts)
                if any(skip in sheet_name.lower() for skip in ["dashboard", "chart", "notes"]):
                    print(f"   ⏩ Skipping display sheet: {sheet_name}")
                    continue

                ws = wb[sheet_name]
                print(f"   📑 Processing sheet: {sheet_name}")
                
                rows_iter = ws.iter_rows(values_only=True)
                
                # 1. Buffer first 15 rows to find a valid header
                header_buffer = []
                try:
                    for _ in range(15):
                        row = next(rows_iter)
                        if any(row): # Only buffer non-empty rows
                            header_buffer.append(row)
                except StopIteration:
                    pass
                
                if not header_buffer:
                    continue

                # Heuristic: Find row with the most text-based column names
                header_idx = 0
                max_strs = 0
                for i, row in enumerate(header_buffer):
                    str_count = sum(1 for cell in row if isinstance(cell, str) and len(str(cell).strip()) > 1)
                    if str_count > max_strs:
                        max_strs = str_count
                        header_idx = i
                
                raw_headers = header_buffer[header_idx]
                headers = [str(h).strip() if h else f"Col_{k}" for k, h in enumerate(raw_headers)]
                
                # 2. Yield from buffer (rows after the header)
                for i in range(header_idx + 1, len(header_buffer)):
                    row_obj = self._make_row(sheet_name, i + 1, header_buffer[i], headers)
                    if row_obj: yield row_obj

                # 3. Stream the rest of the file with an "Early Exit" strategy
                current_row_idx = len(header_buffer) + 1
                consecutive_empty_rows = 0
                
                for row in rows_iter:
                    # Check if the row is effectively empty
                    if not any(cell is not None and str(cell).strip() != "" for cell in row):
                        consecutive_empty_rows += 1
                        # If we see 50 empty rows, assume the data has ended
                        if consecutive_empty_rows > 50:
                            break
                        continue
                    
                    consecutive_empty_rows = 0 # Reset on finding actual data
                    row_obj = self._make_row(sheet_name, current_row_idx, row, headers)
                    if row_obj:
                        yield row_obj
                    
                    current_row_idx += 1
                    
        except Exception as e:
            print(f"❌ Error streaming {self.file_path}: {e}")

    def _make_row(self, sheet, idx, row_tuple, headers):
        """Validates and creates a row object."""
        if not row_tuple:
            return None
            
        # Column 0 is usually the Metric/Label
        metric_name = row_tuple[0]
        values = list(row_tuple[1:])
        
        # Validation: A row is only useful if it has a label and at least one value
        if not metric_name or not any(v is not None for v in values):
            return None
            
        return SpreadsheetRow(
            sheet_name=sheet,
            row_index=idx,
            row_header=metric_name,
            values=values,
            context_tags=headers[1:]
        )